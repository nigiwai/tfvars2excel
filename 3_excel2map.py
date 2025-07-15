import openpyxl
import sys
import json
import os
from dotenv import load_dotenv
from typing import Any, Dict, List, Optional, Tuple, Union

# .envファイルを読み込む
# openpyxl: Excelファイル操作ライブラリ
# dotenv: 設定のための環境変数管理
load_dotenv()
SHEET_PREFIXES = os.getenv("3_SHEET_NAME_PREFIXES").split(",")
DATA_ROW_START = int(os.getenv("3_DATA_ROW_START", "6"))


def pretty_format_tf(value: Any, indent: int = 0) -> str:
    """
    値をTerraform互換の文字列形式でフォーマットする。
    
    引数:
        value: フォーマットする値（辞書、リスト、またはプリミティブ型）
        indent: ネストされた構造のインデントレベル
        
    戻り値:
        Terraform構文でフォーマットされた文字列
        
    注意:
        Terraformの慣例に合わせて2スペースのインデントを使用します。
        ブール値は小文字の'true'/'false'としてフォーマットされます。
    """
    spaces = "  " * indent  # インデントを2スペースに変更
    if isinstance(value, dict):
        if not value:
            return "{}"
        lines = []
        for k, v in value.items():
            key_str = k if k.isidentifier() else json.dumps(k, ensure_ascii=False)
            lines.append(f"{spaces}  {key_str} = {pretty_format_tf(v, indent+1)}")
        return "{\n" + "\n".join(lines) + f"\n{spaces}}}"
    elif isinstance(value, list):
        if not value:
            return "[]"
        lines = []
        for idx, item in enumerate(value):
            line = pretty_format_tf(item, indent + 1)
            if idx < len(value) - 1:
                line += ","
            lines.append(line)
        inner = "\n".join("  " * (indent + 1) + line for line in lines)
        return "[\n" + inner + f"\n{spaces}]"
    else:
        if isinstance(value, str):
            return json.dumps(value, ensure_ascii=False)
        # 修正: bool 値を小文字で出力する
        elif isinstance(value, bool):
            return "true" if value else "false"
        else:
            return str(value)


def pretty_format_map(value: Any, key_string: str = "key", 
                      indent: int = 0) -> str:
    """
    リスト値を生成されたキーを持つマップとしてフォーマットする。
    
    引数:
        value: フォーマットする値
        key_string: 生成されるマップキーの接頭辞
        indent: 現在のインデントレベル
        
    戻り値:
        フォーマットされたマップ文字列
    """
    spaces = "  " * indent
    if isinstance(value, list):
        new_dict = {}
        for i, item in enumerate(value, start=1):
            new_dict[f"{key_string}{i:02d}"] = item
        return pretty_format_tf(new_dict, indent)
    else:
        return pretty_format_tf(value, indent)


def pretty_format_list_object(value: Any, indent: int = 0) -> str:
    """
    各項目に末尾のカンマを付けてオブジェクトのリストをフォーマットする。
    
    引数:
        value: フォーマットするリスト
        indent: 現在のインデントレベル
        
    戻り値:
        末尾にカンマを付けてフォーマットされたリスト文字列
    """
    spaces = "  " * indent
    if isinstance(value, list):
        lines = []
        for i, item in enumerate(value, start=1):
            # 各行末尾に常にカンマを付与
            item_str = pretty_format_tf(item, indent + 1)
            line = f"{spaces}  {item_str},"
            lines.append(line)
        inner = "\n".join(lines)
        return "[\n" + inner + f"\n{spaces}]"
    else:
        return pretty_format_tf(value, indent)


def convert(elem_type: str, val: Any) -> Any:
    """
    文字列値を型指定に基づいて適切な型に変換する。
    
    引数:
        elem_type: 変換先の型（"number", "bool", "list"など）
        val: 変換する値
        
    戻り値:
        変換された値
        
    注意:
        変換に失敗した場合はデータの整合性を保つために元の値を返します。
    """
    if elem_type == "number":
        try:
            return float(val) if "." in val else int(val)
        except (ValueError, TypeError):
            return val
    elif elem_type == "bool":
        return val.lower() == "true"
    elif elem_type == "list":
        if val is None:
            return []
        return [item.strip() for item in val.split(",")]
    elif elem_type == "list(object)":
        if val is None:
            return []
        return val
    elif elem_type == "object":
        if val is None:
            return {}
        return val
    elif elem_type == "map(object)":
        if val is None:
            return {}
        return val
    else:
        return val


def format_value(value: Any, typ: str, field: Optional[str] = None,
                 object_defs: Optional[Dict[str, Any]] = None) -> str:
    """
    terraform.tfvars出力用に型に基づいて値をフォーマットする。
    
    引数:
        value: フォーマットする値
        typ: 型指定（"string", "number", "bool"など）
        field: フィールド名（オプション、マップキー生成に使用）
        object_defs: オブジェクト定義辞書（オプション）
        
    戻り値:
        フォーマットされた値の文字列
        
    例外:
        ValueError: サポートされていない型が指定された場合
    """
    # 空やNoneの場合は type に応じて適切に処理
    if typ in ["string", "number", "bool"]:
        # Noneや空文字の場合はnullで出力
        if not value and value != 0:
            return "null"
        if typ == "bool":
            # 修正: bool値は必ず小文字で出力
            return "true" if value else "false"
        elif typ == "string":
            return f'"{value}"'
        else:  # number
            return str(value)
    elif typ == "list":
        # 空の場合は空リスト
        return json.dumps(value or [], ensure_ascii=False)
    # elif typ == "map":
    #     # 空の場合は {}
    #     return pretty_format_tf(value or {}, 2)
    elif typ == "object":
        # Noneの場合は null
        if value is None:
            return "null"
        return pretty_format_tf(value, 2)
    elif typ == "list(object)":
        # Noneの場合は空リスト
        if value is None:
            return "[]"
        return pretty_format_list_object(value, 2)
    elif typ == "map(object)":
        # Noneの場合は {}
        if value is None:
            return "{}"
        key_str = object_defs.get(field, {}).get("key", field) if object_defs else field
        return pretty_format_map(value, key_str, 2)
    else:
        raise ValueError(f"Unsupported type: {typ}")


def _parse_sheet_metadata(sheet: Any) -> Tuple[List[str], List[str], 
                                               Dict[int, Dict[str, str]], 
                                               List[int]]:
    """
    Excelシートのヘッダー行からメタデータを解析する。
    
    引数:
        sheet: openpyxlワークシートオブジェクト
        
    戻り値:
        以下を含むタプル:
        - headers: 列ヘッダーのリスト
        - column_types: 列の型のリスト
        - object_type_defs: オブジェクト定義の辞書
        - object_field_counts: 列ごとのフィールド数のリスト
        
    例外:
        ValueError: map(object)型に必須の'key'フィールドが不足している場合
    """
    max_col = sheet.max_column
    headers: List[str] = []
    column_types: List[str] = []
    object_type_defs: Dict[int, Dict[str, str]] = {}
    object_field_counts: List[int] = []
    
    for col in range(1, max_col + 1):
        header = sheet.cell(row=1, column=col).value
        headers.append(header)
        col_type = sheet.cell(row=2, column=col).value or "string"
        column_types.append(col_type)
        
        # 3行目からオブジェクト型定義を解析
        if col_type in ["map(object)", "object", "list(object)"]:
            definitions = sheet.cell(row=3, column=col).value
            object_def: Dict[str, str] = {}
            if definitions:
                for line in definitions.splitlines():
                    if ":" in line:
                        elem, elem_type = line.split(":", 1)
                        object_def[elem.strip()] = elem_type.strip()
            
            # map(object)に必須の'key'フィールドを検証
            if col_type == "map(object)" and "key" not in object_def:
                raise ValueError(
                    f"Error: 'key' field is missing for map(object) in {header}"
                )
            object_type_defs[col - 1] = object_def
        
        # 4行目からフィールド数を取得
        object_num = sheet.cell(row=4, column=col).value
        try:
            object_field_counts.append(int(object_num))
        except (ValueError, TypeError):
            object_field_counts.append(0)
    
    return headers, column_types, object_type_defs, object_field_counts


def _merge_object_definitions(headers: List[str], column_types: List[str],
                              object_type_defs: Dict[int, Dict[str, str]]
                              ) -> Dict[str, Dict[str, str]]:
    """
    重複するヘッダーのオブジェクト定義をマージする。
    
    引数:
        headers: 列ヘッダーのリスト
        column_types: 列の型のリスト
        object_type_defs: オブジェクト型定義の辞書
        
    戻り値:
        マージされたオブジェクト定義辞書
    """
    merged_object_defs: Dict[str, Dict[str, str]] = {}
    for i in range(len(headers)):
        if column_types[i] in ["map(object)", "object", "list(object)"]:
            key = headers[i]
            current_def = object_type_defs.get(i, {})
            if key in merged_object_defs:
                merged_object_defs[key].update(current_def)
            else:
                merged_object_defs[key] = current_def
    return merged_object_defs


def _create_reference_map(headers: List[str], column_types: List[str],
                          merged_object_defs: Dict[str, Dict[str, str]]
                          ) -> Dict[str, List[str]]:
    """
    ソースと宛先ヘッダー間の参照マッピングを作成する。
    
    引数:
        headers: 列ヘッダーのリスト
        column_types: 列の型のリスト
        merged_object_defs: マージされたオブジェクト定義
        
    戻り値:
        参照マッピング辞書
    """
    ref_header_map: Dict[str, List[str]] = {}
    for i, hdr in enumerate(headers):
        if (
            column_types[i] in ["map(object)", "object", "list(object)"]
            and hdr in merged_object_defs
        ):
            for field, ftype in merged_object_defs[hdr].items():
                if ftype in ["map(object)", "object", "list(object)", "list"]:
                    if hdr not in ref_header_map:
                        ref_header_map[hdr] = []
                    ref_header_map[hdr].append(field)
    return ref_header_map


def excel_to_tfvars(excel_filepath: str, sheet_title: str,
                    output_filepath: str) -> None:
    """
    Excelシートをterraform.tfvars形式に変換する。
    
    引数:
        excel_filepath: Excelファイルへのパス
        sheet_title: 変換するシート名
        output_filepath: 出力するtfvarsファイルのパス
        
    例外:
        ValueError: データ行のキーがNoneの場合やデータ形式が無効な場合
        
    注意:
        Excelシートの構造:
        - 1行目: ヘッダー
        - 2行目: データ型
        - 3行目: オブジェクト定義（複合型の場合）
        - 4行目: フィールド数
        - 5行目: Description（オプション）
        - 6行目以降: データ行（DATA_ROW_START環境変数で設定可能）
    """
    wb = openpyxl.load_workbook(excel_filepath)
    sheet = wb[sheet_title]

    # シートのメタデータをヘッダー行から解析
    headers, column_types, object_type_defs, object_field_counts = (
        _parse_sheet_metadata(sheet)
    )
    
    # 重複ヘッダーの定義をマージ
    merged_object_defs = _merge_object_definitions(
        headers, column_types, object_type_defs
    )
    
    # ネストされたオブジェクトの関係のための参照マッピング作成
    ref_header_map = _create_reference_map(
        headers, column_types, merged_object_defs
    )

    # ヘッダーから型へのマッピング作成（最初の列はキーなので除外）
    header_type_dict: Dict[str, str] = {
        headers[i]: column_types[i] for i in range(1, len(headers))
    }

    # データ行の処理
    tfvars_data_map: Dict[str, Dict[str, Any]] = {}
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=DATA_ROW_START, values_only=True), start=DATA_ROW_START
    ):
        row_key = row[0]
        if row_key is None:
            raise ValueError(f"Error: Key is None in row {row_idx}")
        
        row_values: Dict[str, Any] = {}
        for idx in range(1, len(headers)):
            cell_value = row[idx]
            col_type = column_types[idx]
            header = headers[idx]
            object_num = object_field_counts[idx]
            # 型に応じた変換処理
            if cell_value is not None or cell_value == "":
                if col_type == "string":
                    converted_value = str(cell_value)
                elif col_type == "number":
                    try:
                        converted_value = (
                            float(cell_value)
                            if "." in str(cell_value)
                            else int(cell_value)
                        )
                    except:
                        converted_value = cell_value

                # ブール値の厳密な検証
                elif col_type == "bool":
                    if cell_value not in ("true", "false"):
                        raise ValueError(
                            f"Error in sheet '{sheet.title}' for key '{row_key}', "
                            f"type:{col_type} '{header}' expects 'true' or 'false' "
                            f"but got '{cell_value}'"
                        )
                    converted_value = cell_value.strip().lower() == "true"
                elif col_type == "list":
                    converted_value = (
                        cell_value.splitlines()
                        if (isinstance(cell_value, str) and "\n" in cell_value)
                        else [cell_value]
                    )
                elif col_type in ["map(object)", "object", "list(object)"]:
                    lines = str(cell_value).splitlines()
                    if header in merged_object_defs:
                        objects = []
                        for line in lines:
                            values_list = [p.strip() for p in line.split(":")]
                            if len(values_list) != object_field_counts[idx]:
                                raise ValueError(
                                    f"Error in sheet '{sheet.title}' for key "
                                    f"'{row_key}', type:{col_type} '{header}' "
                                    f"expects {object_field_counts[idx]} elements "
                                    f"but got {len(values_list)} in line: {line}"
                                )
                            obj = {}
                            for i, key_elem in enumerate(
                                merged_object_defs[header].keys()
                            ):
                                if col_type == "map(object)" and i == 0:
                                    continue
                                idx_offset = i - 1 if col_type == "map(object)" else i
                                raw = (
                                    values_list[idx_offset]
                                    if idx_offset < len(values_list)
                                    else None
                                )
                                elem_type = merged_object_defs[header][key_elem]
                                obj[key_elem] = convert(elem_type, raw)
                            objects.append(obj)
                        if col_type == "object":
                            converted_value = objects[0] if objects else None
                        else:
                            converted_value = objects
                    else:
                        converted_value = cell_value
                else:
                    converted_value = cell_value
                row_values[header] = converted_value
            else:
                if col_type in ["map(object)", "object"]:
                    row_values[header] = {}
                elif col_type in ["list", "list(object)"]:
                    row_values[header] = []
                else:
                    row_values[header] = None
        # 変更: 参照先の値を参照元のmapsにネストして結合する
        for src_hdr, dest_list in ref_header_map.items():
            for dest_hdr in dest_list:
                if src_hdr in {h for h in headers} and dest_hdr in {h for h in headers}:
                    if (
                        src_hdr in row_values
                        and dest_hdr in row_values
                        and row_values[dest_hdr] is not None
                    ):
                        src_val = row_values[src_hdr]
                        if not isinstance(src_val, list):
                            src_val = [src_val]
                        dest_val = row_values[dest_hdr]
                        if isinstance(dest_val, list):
                            dest_objs = dest_val
                        else:
                            dest_objs = [dest_val]
                        if not dest_objs:
                            continue
                        for obj_index, obj in enumerate(src_val):
                            if obj is None or not isinstance(obj, dict):
                                obj = {}
                                src_val[obj_index] = obj
                            counter = 1
                            dest_def = merged_object_defs.get(dest_hdr, {})
                            dest_prefix = dest_def.get("key", dest_hdr)
                            for d_obj in dest_objs:
                                dest_field_type = header_type_dict.get(dest_hdr)
                                if dest_field_type == "list":
                                    if dest_hdr not in obj or not isinstance(
                                        obj[dest_hdr], list
                                    ):
                                        obj[dest_hdr] = []
                                    obj[dest_hdr].append(d_obj)
                                elif dest_field_type == "object":
                                    if dest_hdr not in obj or not isinstance(
                                        obj[dest_hdr], dict
                                    ):
                                        obj[dest_hdr] = {}
                                    obj[dest_hdr] = d_obj
                                elif dest_field_type in ["map(object)"]:
                                    if dest_hdr not in obj or not isinstance(
                                        obj[dest_hdr], dict
                                    ):
                                        obj[dest_hdr] = {}
                                    sub_key = f"{dest_prefix}{counter:02d}"
                                    obj[dest_hdr][sub_key] = d_obj
                                elif dest_field_type == "list(object)":
                                    if dest_hdr not in obj or not isinstance(
                                        obj[dest_hdr], list
                                    ):
                                        obj[dest_hdr] = []
                                    obj[dest_hdr].append(d_obj)
                                else:
                                    raise ValueError(
                                        f"Error: Unsupported type '{dest_field_type}' for '{dest_hdr}'"
                                    )
                                counter += 1
                            row_values[src_hdr] = src_val
                    row_values.pop(dest_hdr, None)
        tfvars_data_map[row_key] = row_values
    # 各ヘッダーの型情報をマッピング（先頭列以外）

    # tfvarsファイルとして出力（Terraformの各typeをシンプルに処理）
    mode = "a" if os.path.exists(output_filepath) else "w"
    with open(output_filepath, mode, encoding="utf-8", newline="\n") as tfvars_file:
        tfvars_file.write(f"{sheet.title} = {{\n")

        # シート内の各キーごとにブロックを出力
        for key, data in tfvars_data_map.items():
            tfvars_file.write(f'  "{key}" = {{\n')

            for field, value in data.items():
                typ = header_type_dict.get(field)
                val_str = format_value(value, typ, field, merged_object_defs)
                tfvars_file.write(f"    {field} = {val_str}\n")

            tfvars_file.write("  },\n")
        tfvars_file.write("}\n")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python 3_excel2map.py <excel_filepath>")
        sys.exit(1)

    excel_file_path = sys.argv[1]

    # 出力ディレクトリ構造の作成
    excel_file_name = os.path.splitext(os.path.basename(excel_file_path))[0]
    output_folder = os.path.join("output", excel_file_name)
    os.makedirs(output_folder, exist_ok=True)
    output_tfvars_file = os.path.join(output_folder, "terraform.tfvars")

    # 設定されたプレフィックスに一致するシートを処理
    wb = openpyxl.load_workbook(excel_file_path)
    for sheet_name in wb.sheetnames:
        if any(sheet_name.startswith(prefix) for prefix in SHEET_PREFIXES):
            print(f"Converting {sheet_name} sheet to terraform.tfvars")
            excel_to_tfvars(excel_file_path, sheet_name, output_tfvars_file)
