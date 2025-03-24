import json
import openpyxl
import sys
import warnings
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from dotenv import load_dotenv

load_dotenv()  # 環境変数を読み込み

# tfvarsファイルを読み込む関数
def load_tfvars(filepath):
    with open(filepath, 'r', encoding='utf-8') as file:
        lines = file.readlines()
    tfvars = {}
    current_key = None
    current_value = []
    for line in lines:
        if '=' in line:
            if current_key:
                tfvars[current_key] = ''.join(current_value).split('#')[0].strip()
            key, value = line.split('=', 1)
            current_key = key.strip()
            current_value = [value.strip()]
        else:
            current_value.append(line.strip())
    if current_key:
        tfvars[current_key] = ''.join(current_value).split('#')[0].strip()
    return tfvars

# tfvarsの値を整形する関数
def format_tfvars_value(value):
    # 日本語コメント: 項目が空(null, {}, [], 空文字)の場合は空文字を返す
    value = value.strip()
    if value in ['null', '{}', '[]', '']:
        return ''
    # 日本語コメント: ダブルクォートを除去
    value = value.strip('"')
    if value.startswith('[') and value.endswith(']'):
        value = value[1:-1]  # []を除去
        value = value.replace('"', '')  # "を除去
        value = value.replace(', ', '\n')  # ,を改行に置換
        value = value.replace(',', '\n')  # 改行されている場合の,を改行に置換
        value = value.replace('\n ', '\n')  # 改行後の先頭の空白を除去
        value = '\n'.join([v.strip() for v in value.split('\n') if v.strip()])  # 空白行を除去
    return value

# tfvarsの値の型を判定する関数
def determine_type(value):
    # 日本語コメント: 項目が空(null, {}, [], 空文字)の場合はunknown型とする
    value = value.strip()
    if value in ['', 'null', '{}', '[]']:
        return 'unknown'
    if value.startswith('['):
        return 'list'
    elif value.startswith('{'):
        return 'map'
    elif value.startswith('"') and value.endswith('"'):
        return 'string'
    elif value.isdigit() or (value.startswith('-') and value[1:].isdigit()):
        return 'number'
    elif value in ['true', 'false']:
        return 'bool'
    return 'unknown'

def contains_japanese(text):
    # 日本語判定: 全角ひらがなカタカナ漢字チェック
    return bool(re.search(r'[ぁ-んァ-ン一-龯]', text))

def should_skip(item):
    ban_words = (os.getenv("2_BAN_WORDS") or "").split(",")
    # ban_wordsが空にならないようstrip
    ban_words = [bw.strip() for bw in ban_words if bw.strip()]
    # BAN_WORDSに含まれる単語、または日本語を含むならスキップ
    if any(bw in item for bw in ban_words):
        return True
    if contains_japanese(item):
        return True
    return False

# Excelファイルを更新する関数
def update_excel(tfvars, excel_filepath):
    if not excel_filepath.endswith('.xlsx'):
        raise ValueError("Only .xlsx files are supported")
    wb = load_workbook(excel_filepath)
    ws = wb["ヒアリングシート"]

    unmatched_keys = set(tfvars.keys())

    excel_items = []
    for row in ws.iter_rows(min_row=2, max_col=6, max_row=ws.max_row):
        a_value = row[0].value
        fill = row[0].fill
        if a_value in tfvars and (fill is None or fill == PatternFill()):
            formatted_value = format_tfvars_value(tfvars[a_value])
            row[5].value = formatted_value
            row[1].value = determine_type(tfvars[a_value])
            if formatted_value == '':
                row[3].value = 'false'
            unmatched_keys.discard(a_value)

    for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
        if row[0].value:
            excel_items.append(row[0].value)

    excel_only = set(excel_items) - set(tfvars.keys())

    tfvars_only = {k for k in unmatched_keys if not should_skip(k)}
    excel_only_filtered = {k for k in excel_only if not should_skip(k)}
    

    if tfvars_only:
        print("-----tfvars only:")
        for key in tfvars_only:
            print(key)

    if excel_only_filtered:
        print("-----excel only:")
        for key in excel_only_filtered:
            print(key)

    # Excelファイルを保存
    wb.save(excel_filepath)

    # Excelファイルを開く
    os.system(f'start excel "{excel_filepath}"')

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python tfvars2excel.py <excel_filepath> <tfvars_filepath>")
        sys.exit(1)

    excel_filepath = sys.argv[1]
    
    filename = os.path.splitext(os.path.basename(excel_filepath))[0]
    if sys.argv[2]:
        tfvars_filepath = sys.argv[2]
    else:
        tfvars_filepath = f"./output/{filename}/terraform.tfvars"

    tfvars = load_tfvars(tfvars_filepath)
    update_excel(tfvars, excel_filepath)
