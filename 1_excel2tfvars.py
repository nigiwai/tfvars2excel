import sys
import os
import openpyxl
from openpyxl import load_workbook
from dotenv import load_dotenv  # .envを読み込むためのライブラリ
load_dotenv()  # .envを読み込み

SHEET_PREFIXES = os.getenv("1_SHEET_NAME_PREFIXES", "")  # .envからシート名のプレフィックスを取得

def read_excel(file_path):
    valid_extensions = [".xlsx", ".xlsm", ".xltx", ".xltm"]
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in valid_extensions:
        raise ValueError(f"Unsupported file format: {ext}; please use xlsx, xlsm, xltx, or xltm.")

    wb = load_workbook(file_path)
    sheets_data = {}
    for sheet_name in wb.sheetnames:
        # シート名が指定したプレフィックスで始まるかチェック
        if sheet_name.startswith(SHEET_PREFIXES):
            sheet = wb[sheet_name]
            # values_only=Trueだとセルオブジェクトが得られないので、シートはそのまま保持
            data = []
            for row in sheet.iter_rows(min_row=3, values_only=True):
                data.append(row)
            sheets_data[sheet_name] = (data, sheet)
    return sheets_data

def validate_value(value, value_type, max_length_or_limit, allow_empty):
    # 値が空の場合の基本チェック
    if value is None:
        if not allow_empty:
            raise ValueError("Empty value is not allowed")
        return ""

    if value_type == "string":
        if max_length_or_limit and len(value) > max_length_or_limit:
            raise ValueError(
                f"String length exceeds the maximum length of {max_length_or_limit}"
            )
    elif value_type == "number":
        numeric_value = float(value)
        if max_length_or_limit and numeric_value > float(max_length_or_limit):
            raise ValueError(f"Number exceeds the maximum limit of {max_length_or_limit}")
        # 整数なら整数文字列、少数なら少数文字列を返す
        return str(int(numeric_value)) if numeric_value.is_integer() else str(numeric_value)
    elif value_type == "bool":
        if value not in ["true", "false"]:
            raise ValueError("Boolean value must be 'true' or 'false'")
    return value

def generate_tfvars(sheets_data, output_file):
    # TFVARSファイルを書き出す関数
    with open(output_file, "w", encoding="utf-8", newline="\n") as f:
        for sheet_name, (data, sheet) in sheets_data.items():
            for index, row in enumerate(data):
                # シートのA列(変数名)を取得
                cell = sheet[f"A{index + 3}"]
                # 色なしのセルだけを処理
                if cell.fill.start_color.index == "00000000":
                    var_name = row[0]
                    var_type = row[1]
                    max_length_or_limit = row[2]
                    allow_empty = row[3]
                    value = row[5]
                    # 値が空の場合はそのままにしてvalidate_valueへ渡す
                    try:
                        validated_value = validate_value(
                            value, var_type, max_length_or_limit, allow_empty
                        )
                    except ValueError as e:
                        # エラー時は終了
                        print(f"Error validating value for {var_name}: {e}")
                        sys.exit(1)
                
                    # 空白の場合の出力をシンプルに
                    if not validated_value:
                        # string/number/boolの場合はnullにする
                        if var_type in ["string", "number", "bool"]:
                            f.write(f"{var_name} = null\n")
                        elif var_type == "list":
                            f.write(f"{var_name} = []\n")
                        elif var_type == "map":
                            f.write(f"{var_name} = {{}}\n")
                        else:
                            f.write(f"{var_name} = null\n")
                    else:
                        # 各Terraform型ごとに書き出し
                        if var_type == "string":
                            f.write(f'{var_name} = "{validated_value}"\n')
                        elif var_type == "list":
                            values = validated_value.split("\n")
                            f.write(f"{var_name} = [\n")
                            for i, val in enumerate(values):
                                # 最後の要素にカンマを付与しない
                                if i < len(values) - 1:
                                    f.write(f'    "{val}",\n')
                                else:
                                    f.write(f'    "{val}"\n')
                            f.write("]\n")
                        elif var_type == "map":
                            values = validated_value.split("\n")
                            f.write(f"{var_name} = {{\n")
                            for i, val in enumerate(values):
                                key_values = val.split(":")
                                f.write(f'    "key{i+1:03d}" = {{\n')
                                for j, kv in enumerate(key_values):
                                    f.write(f'        "value{j+1:03d}" = "{kv}",\n')
                                f.write("    },\n")
                            f.write("}\n")
                        elif var_type == "number":
                            f.write(f"{var_name} = {validated_value}\n")
                        elif var_type == "bool":
                            # boolは小文字にして書き出す
                            f.write(f"{var_name} = {str(validated_value).strip().lower()}\n")
                        else:
                            f.write(f"{var_name} = null\n")

if __name__ == "__main__":
    # コマンドライン引数チェック
    if len(sys.argv) != 2:
        print("Usage: python excel2tfvars.py <excel_filepath>")
        sys.exit(1)

    excel_file_path = sys.argv[1]

    # Excelファイル名からフォルダ名を作成
    excel_file_name = os.path.splitext(os.path.basename(excel_file_path))[0]
    output_folder = os.path.join("output", excel_file_name)
    os.makedirs(output_folder, exist_ok=True)
    output_tfvars_file = os.path.join(output_folder, "terraform.tfvars")

    sheets_data = read_excel(excel_file_path)
    if not sheets_data:
        print(f"No sheets found starting with prefix: {SHEET_PREFIXES}")
        sys.exit(1)
    generate_tfvars(sheets_data, output_tfvars_file)
